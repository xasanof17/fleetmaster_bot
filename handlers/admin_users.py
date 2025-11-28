from aiogram import Router, F
from aiogram.types import (
    Message,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from aiogram.filters import StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.filters import Command
from config import settings
from services.access_control import (
    access_storage,
    AccessStatus,
    AccessRequest,
    UserRole,
)
from utils.logger import get_logger
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

router = Router()
logger = get_logger("handlers.admin_users")

ADMINS = set(settings.ADMINS or [])


# ================================================================
# HELPER: ADMIN CHECK
# ================================================================
def is_admin(user_id: int) -> bool:
    return user_id in ADMINS


# ================================================================
# ADMIN PANEL MAIN MENU
# ================================================================
def admin_menu_keyboard():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📋 Pending users", callback_data="admin:pending")],
            [InlineKeyboardButton(text="✅ Approved users", callback_data="admin:approved")],
            [InlineKeyboardButton(text="❌ Denied users", callback_data="admin:denied")],
            [InlineKeyboardButton(text="🔍 Search user", callback_data="admin:search")],
            [InlineKeyboardButton(text="📤 Export to Excel", callback_data="admin:export")],
        ]
    )


@router.message(Command("admin"))
async def cmd_admin(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ You do not have permission.")
        return

    await message.answer(
        "👨‍💼 Admin Panel – Manage Users:", 
        reply_markup=admin_menu_keyboard()
    )


# ================================================================
# LIST USERS BY STATUS
# ================================================================
def _build_user_list(status: AccessStatus):
    base_dir = access_storage.base_dir / status.value
    users = []

    for file in base_dir.iterdir():
        try:
            req = access_storage.get(int(file.stem))
            if req:
                users.append(req)
        except:
            continue

    return users


def _format_user_line(req: AccessRequest):
    return (
        f"{req.full_name} | {req.role}\n"
        f"{req.gmail} | {req.phone}\n"
        f"ID: {req.tg_id}\n"
        f"{'-'*30}\n"
    )


@router.callback_query(F.data.startswith("admin:pending"))
async def admin_pending(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return

    users = _build_user_list(AccessStatus.PENDING)

    if not users:
        await callback.message.edit_text("📋 No pending users.")
        return

    text = "📋 <b>Pending Users</b>\n\n"
    for req in users:
        text += _format_user_line(req)

    await callback.message.edit_text(text, parse_mode="HTML")


@router.callback_query(F.data.startswith("admin:approved"))
async def admin_approved(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return

    users = _build_user_list(AccessStatus.APPROVED)

    if not users:
        await callback.message.edit_text("✅ No approved users.")
        return

    text = "✅ <b>Approved Users</b>\n\n"
    for req in users:
        text += _format_user_line(req)

    await callback.message.edit_text(text, parse_mode="HTML")


@router.callback_query(F.data.startswith("admin:denied"))
async def admin_denied(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return

    users = _build_user_list(AccessStatus.DENIED)

    if not users:
        await callback.message.edit_text("❌ No denied users.")
        return

    text = "❌ <b>Denied Users</b>\n\n"
    for req in users:
        text += _format_user_line(req)

    await callback.message.edit_text(text, parse_mode="HTML")


# ================================================================
# SEARCH USER MODE
# ================================================================
@router.callback_query(F.data == "admin:search")
async def admin_search_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("🔍 Enter name, Gmail, or role to search:")
    await state.set_state("admin_search")


@router.message( StateFilter("admin_search"))
async def admin_search_process(message: Message, state: FSMContext):
    query = message.text.lower().strip()
    await state.clear()

    found = []

    for status in AccessStatus:
        users = _build_user_list(status)
        for req in users:
            if (
                query in req.full_name.lower()
                or query in req.gmail.lower()
                or query in req.role.lower()
                or query in req.phone.lower()
            ):
                found.append(req)

    if not found:
        await message.answer("❌ No users found.")
        return

    text = "🔍 <b>Search Results</b>\n\n"
    for req in found:
        text += _format_user_line(req)

    await message.answer(text, parse_mode="HTML")


# ================================================================
# EXPORT FULL USER LIST TO EXCEL
# ================================================================
@router.callback_query(F.data == "admin:export")
async def admin_export_users(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return

    file_path = "users_export.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    # Columns
    headers = [
        "Telegram ID",
        "Full Name",
        "Gmail",
        "Phone",
        "Telegram Username",
        "Role",
        "Status",
        "Created At",
        "Approved By",
    ]

    # Header styles
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", fill_type="solid")

    # Fill data
    row = 2
    for status in AccessStatus:
        users = _build_user_list(status)
        for req in users:
            ws.cell(row=row, column=1, value=req.tg_id)
            ws.cell(row=row, column=2, value=req.full_name)
            ws.cell(row=row, column=3, value=req.gmail)
            ws.cell(row=row, column=4, value=req.phone)
            ws.cell(row=row, column=5, value=req.telegram_username)
            ws.cell(row=row, column=6, value=req.role)
            ws.cell(row=row, column=7, value=req.status.value)
            ws.cell(row=row, column=8, value=req.created_at)
            ws.cell(row=row, column=9, value=req.approved_by)
            row += 1

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3

    wb.save(file_path)

    await callback.message.answer_document(
        document=file_path,
        caption="📤 Exported Users Excel File"
    )

    os.remove(file_path)
    await callback.answer("Exported successfully.")
